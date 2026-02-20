"""
Copyright 2023, 2024, 2025, 2026 Jack Consoli.  All rights reserved.

Licensed under the Apache License, Version 2.0 (the "License"); you may not use this file except in compliance with
the License. You may also obtain a copy of the License at https://www.apache.org/licenses/LICENSE-2.0

Unless required by applicable law or agreed to in writing, software distributed under the License is distributed on an
"AS IS" BASIS, WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied. See the License for the specific
language governing permissions and limitations under the License.

The license is free for single customer use (internal applications). Use of this module in the production,
redistribution, or service delivery for commerce requires an additional license. Contact jack_consoli@yahoo.com for
details.

:mod:`brcdapi.excel_fonts` - Contains miscellaneous font, fill, and alignment definitions for the openpyxl library

**Public Methods & Data**

+-----------------------+-------------------------------------------------------------------------------------------+
| Method                | Description                                                                               |
+=======================+===========================================================================================+
| font_type             | Returns the font definition tuple for the openpyxl libraries. See _font_d keys.           |
+-----------------------+-------------------------------------------------------------------------------------------+
| fill_type             | Returns the font definition tuple for the openpyxl libraries. See _fill_d keys.           |
+-----------------------+-------------------------------------------------------------------------------------------+
| border_type           | Returns the border definition tuple for the openpyxl libraries. See _border_d keys.       |
+-----------------------+-------------------------------------------------------------------------------------------+
| align_type            | Returns the alignment definition tuple for the openpyxl libraries. See _align_d keys.     |
+-----------------------+-------------------------------------------------------------------------------------------+

**Version Control**

+-----------+---------------+---------------------------------------------------------------------------------------+
| Version   | Last Edit     | Description                                                                           |
+===========+===============+=======================================================================================+
| 4.0.0     | 04 Aug 2023   | Re-Launch                                                                             |
+-----------+---------------+---------------------------------------------------------------------------------------+
| 4.0.1     | 06 Mar 2024   | Documentation updates only.                                                           |
+-----------+---------------+---------------------------------------------------------------------------------------+
| 4.0.2     | 20 Oct 2024   | Added fonts and fill types                                                            |
+-----------+---------------+---------------------------------------------------------------------------------------+
| 4.0.3     | 25 Aug 2025   | Updated email address in __email__ only.                                              |
+-----------+---------------+---------------------------------------------------------------------------------------+
| 4.0.4     | 19 Oct 2025   | Updated comments only.                                                                |
+-----------+---------------+---------------------------------------------------------------------------------------+
| 4.0.5     | 20 Feb 2026   | Updated copyright notice.                                                             |
+-----------+---------------+---------------------------------------------------------------------------------------+
"""
__author__ = 'Jack Consoli'
__copyright__ = 'Copyright 2024, 2025, 2026 Jack Consoli'
__date__ = '20 Feb 2026'
__license__ = 'Apache License, Version 2.0'
__email__ = 'jack_consoli@yahoo.com'
__maintainer__ = 'Jack Consoli'
__status__ = 'Released'
__version__ = '4.0.5'

import brcdapi.log as brcdapi_log
import openpyxl.styles as xl_styles

#################################################################
#   Font Types                                                  #
#                                                               #
#   If you add a font type, remember to add it to _font_d       #
#################################################################
_cli_font = xl_styles.Font(
    name='Courier New',
    size=11,
    bold=False,
    italic=False,
    vertAlign=None,
    underline='none',
    strike=False,
    color='FF000000')
_std_font = xl_styles.Font(
    name='Calibri',
    size=11,
    bold=False,
    italic=False,
    vertAlign=None,
    underline='none',
    strike=False,
    color='FF000000')
_h1_font = xl_styles.Font(
    name='Calibri',
    size=14,
    bold=True,
    italic=False,
    vertAlign=None,
    underline='none',
    strike=False,
    color='FF000000')
_h2_font = xl_styles.Font(
    name='Calibri',
    size=12,
    bold=True,
    italic=False,
    vertAlign=None,
    underline='none',
    strike=False,
    color='FF000000')
_bold_font = xl_styles.Font(
    name='Calibri',
    size=11,
    bold=True,
    italic=False,
    vertAlign=None,
    underline='none',
    strike=False,
    color='FF000000')
_italic_font = xl_styles.Font(
    name='Calibri',
    size=11,
    bold=False,
    italic=True,
    vertAlign=None,
    underline='none',
    strike=False,
    color='FF000000')
_warn_font = xl_styles.Font(
    name='Calibri',
    size=11,
    bold=True,
    italic=False,
    vertAlign=None,
    underline='none',
    strike=False,
    color='FF8000')
_error_font = xl_styles.Font(
    name='Calibri',
    size=11,
    bold=True,
    italic=False,
    vertAlign=None,
    underline='none',
    strike=False,
    color='FF0000')
_link_font = xl_styles.Font(
    name='Calibri',
    size=11,
    bold=False,
    italic=False,
    vertAlign=None,
    underline='single',
    strike=False,
    color='3336FF')
_white_bold_font = xl_styles.Font(
    name='Calibri',
    size=11,
    bold=True,
    italic=False,
    vertAlign=None,
    underline='none',
    strike=False,
    color='FFFFFFFF')
_gray_font = xl_styles.Font(
    name='Calibri',
    size=11,
    bold=False,
    italic=False,
    vertAlign=None,
    underline='none',
    strike=False,
    color='FF808080')
_font_d = dict(
    cli=_cli_font,
    std=_std_font,
    hdr_1=_h1_font,
    hdr_2=_h2_font,
    bold=_bold_font,
    italic=_italic_font,
    warn=_warn_font,
    error=_error_font,
    link=_link_font,
    white_bold=_white_bold_font,
    gray=_gray_font
)

#################################################################
#   Fill Types                                                  #
#                                                               #
#   If you add a fill type, remember to add it to _fill_d       #
#################################################################
_lightblue_fill = xl_styles.PatternFill(  # Fill used in reports
    fill_type='solid',
    start_color='FFCCE5FF',
    end_color='FFCCE5FF',
)
_lightgreen_fill = xl_styles.PatternFill(  # Fill used in reports
    fill_type='solid',
    start_color='FFE2EFDA',
    end_color='FFE2EFDA',
)
_lightgold_fill = xl_styles.PatternFill(  # Fill used in reports
    fill_type='solid',
    start_color='FFFFF2CC',
    end_color='FFFFF2CC',
)
_lightred_fill = xl_styles.PatternFill(  # Fill used in reports
    fill_type='solid',
    start_color='FFFFE7E7',
    end_color='FFFFE7E7',
)
_yellow_fill = xl_styles.PatternFill(  # Fill used in reports
    fill_type='solid',
    start_color='FFFFFF00',
    end_color='FFFFFF00',
)
_config_asic_0_fill = xl_styles.PatternFill(  # orange color used in switch configuration workbook for ASIC 0
    fill_type='solid',
    start_color='FFFCD5B4',
    end_color='FFFCD5B4',
)
_config_asic_1_fill = xl_styles.PatternFill(  # grey-blue color used in switch configuration workbook for ASIC 1
    fill_type='solid',
    start_color='FFDCE6F1',
    end_color='FFDCE6F1',
)
_config_slot_fill = xl_styles.PatternFill(  # Fill used in reports
    fill_type='solid',
    start_color='FFC00000',
    end_color='FFC00000',
)
_fill_d = dict(
    lightblue=_lightblue_fill,
    config_asic_0=_config_asic_0_fill,
    config_asic_1=_config_asic_1_fill,
    config_slot=_config_slot_fill,
    lightgreen=_lightgreen_fill,
    lightgold=_lightgold_fill,
    lightred=_lightred_fill,
    yellow=_yellow_fill,
)

#################################################################
#   Border Types                                                #
#                                                               #
#   If you add a border type, remember to add it to _border_d #
#################################################################
_thin_border = xl_styles.Border(
    left=xl_styles.Side(border_style='thin', color='FF000000'),
    right=xl_styles.Side(border_style='thin', color='FF000000'),
    top=xl_styles.Side(border_style='thin', color='FF000000'),
    bottom=xl_styles.Side(border_style='thin', color='FF000000'),
    # diagonal=xl_styles.Side(border_style=None,color='FF000000'),
    # diagonal_direction=0,outline=xl_styles.Side(border_style=None,color='FF000000'),
    # vertical=xl_styles.Side(border_style=None,color='FF000000'),
    # horizontal=xl_styles.Side(border_style=None,color='FF000000')
)
_border_d = dict(
    thin=_thin_border,
)

#################################################################
#   Align Types                                                 #
#                                                               #
#   If you add an align type, remember to add it to _align_d  #
#################################################################
_wrap_center_alignment = xl_styles.Alignment(
    horizontal='center',
    vertical='top',
    text_rotation=0,
    wrap_text=True,
    shrink_to_fit=False,
    indent=0)
_wrap_right_alignment = xl_styles.Alignment(
    horizontal='right',
    vertical='top',
    text_rotation=0,
    wrap_text=True,
    shrink_to_fit=False,
    indent=0)
_wrap_alignment = xl_styles.Alignment(
    horizontal='general',
    vertical='top',
    text_rotation=0,
    wrap_text=True,
    shrink_to_fit=False,
    indent=0)
_wrap_vert_center_alignment = xl_styles.Alignment(
    horizontal='center',
    vertical='top',
    text_rotation=90,
    wrap_text=True,
    shrink_to_fit=False,
    indent=0)
_align_d = dict(
    wrap_center=_wrap_center_alignment,
    wrap=_wrap_alignment,
    wrap_vert_center=_wrap_vert_center_alignment,
    wrap_right=_wrap_right_alignment,
)


def font_type(x):
    """Returns the font definition tuple for the openpyxl libraries

    :param x: Font type listed in _font_d
    :type x: str
    :return: Font definitions for openpyxl library
    :rtype: tuple
    """
    global _font_d, _std_font

    try:
        return _font_d[x] if isinstance(x, str) else None
    except KeyError:
        brcdapi_log.exception('Unknown font type: ' + x, True)
    return _std_font


def fill_type(x):
    """Returns the fill definition tuple for the openpyxl libraries

    :param x: Fill type listed in _fill_d
    :type x: str
    :return: Fill definitions for openpyxl library
    :rtype: tuple
    """
    global _fill_d

    try:
        return _fill_d[x] if isinstance(x, str) else None
    except KeyError:
        brcdapi_log.exception('Unknown fill type: ' + x, True)
    return None


def border_type(x):
    """Returns the border definition tuple for the openpyxl libraries

    :param x: Border type listed in _border_d
    :type x: str
    :return: Border definitions for openpyxl library
    :rtype: tuple
    """
    global _border_d

    try:
        return _border_d[x] if isinstance(x, str) else None
    except KeyError:
        brcdapi_log.exception('Unknown border type: ' + x, True)
    return None


def align_type(x):
    """Returns the alignment definition tuple for the openpyxl libraries

    :param x: Alignment type listed in _align_d
    :type x: str
    :return: Alignment definitions for openpyxl library
    :rtype: tuple
    """
    global _align_d

    try:
        return _align_d[x] if isinstance(x, str) else None
    except KeyError:
        brcdapi_log.exception('Unknown align type: ' + x, True)
    return None
